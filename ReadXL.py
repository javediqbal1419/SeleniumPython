import  openpyxl

path="C:/Users/javed iqbal/Box Sync/selenium/ReadXL.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.get_sheet_by_name("user")
rows=sheet.max_row
cols=sheet.max_column

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end="    ")
    print()

